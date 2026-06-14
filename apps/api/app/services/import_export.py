from __future__ import annotations

import io
import re
import uuid
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select, func
from sqlalchemy.orm import Session

from app.models.test_case import TestCase
from app.models.test_step import TestStep
from app.models.user import User
from app.schemas.import_export import (
    DuplicateAction,
    DuplicateRow,
    ImportError,
    ImportExecuteRequest,
    ImportResult,
    ImportStepRow,
    ParseResult,
    TestCaseImportRow,
)

MAX_IMPORT_ROWS = 500

COL = {
    "display_id":        "TC ID",
    "title":             "Title",
    "type":              "Type",
    "severity":          "Severity",
    "status":            "Status",
    "description":       "Description",
    "preconditions":     "Preconditions",
    "step_actions":      "Step Actions",
    "expected_result":   "Expected Result",
    "notes":             "Notes",
    "automation_status": "Automation Status",
    "environment":       "Environment",
    "estimated_time":    "Estimated Time",
    "tags":              "Tags",
    "requirement_id":    "Requirement ID",
    "assigned_to_name":  "Assigned To",
}

EXPORT_HEADER = list(COL.values())

REQUIRED_COLUMNS = {
    COL["title"],
    COL["type"],
    COL["severity"],
    COL["expected_result"],
    COL["step_actions"],
}

VALID_TYPES = {"Functional", "Regression", "Smoke", "Integration", "UI", "Performance", "Security"}
VALID_SEVERITIES = {"Blocker", "Critical", "Major", "Minor"}
VALID_STATUSES = {"Draft", "Ready", "In Review", "Approved", "Obsolete"}
VALID_ENVIRONMENTS = {"Staging", "Production", "UAT", "Development", None}
VALID_AUTOMATION_STATUSES = {"Manual", "Automated", "Candidate to Automate", None}


def _numbered_list(items: list[str | None]) -> str | None:
    lines = []
    for i, item in enumerate(items, start=1):
        if item:
            lines.append(f"{i}. {item}")
    return "\n".join(lines) if lines else None


def _parse_numbered_list(text: str | None) -> list[str | None]:
    if not text:
        return []
    lines = [line.strip() for line in text.strip().splitlines() if line.strip()]
    result: dict[int, str] = {}
    max_n = 0
    for line in lines:
        match = re.match(r"^(\d+)\.\s*(.*)$", line)
        if match:
            n = int(match.group(1))
            result[n] = match.group(2).strip()
            if n > max_n:
                max_n = n
        else:
            pass
    if not result:
        return [text.strip()] if text.strip() else []
    return [result.get(i) for i in range(1, max_n + 1)]


def _parse_steps_with_details(text: str | None) -> list[dict]:
    if not text:
        return []
    lines = text.strip().splitlines()
    steps = []
    current_step = None
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        match = re.match(r"^(\d+)\.\s*(.*)$", stripped)
        if match:
            step_num = int(match.group(1))
            action = match.group(2).strip()
            current_step = {
                "step_number": step_num,
                "action": action,
                "test_data": None,
                "expected_result": None
            }
            steps.append(current_step)
        elif current_step:
            if stripped.startswith("Test Data:"):
                current_step["test_data"] = stripped[len("Test Data:"):].strip()
            elif stripped.startswith("Expected:"):
                current_step["expected_result"] = stripped[len("Expected:"):].strip()
            else:
                current_step["action"] += "\n" + stripped
    return steps


def _sanitize_sheet_name(name: str, existing_names: set[str]) -> str:
    cleaned = re.sub(r"[\\/\?\*:\[\]]", "", name).strip(" '\"")
    if not cleaned:
        cleaned = "Module"
    base = cleaned[:31]
    
    # Check collision (case-insensitive)
    if base.lower() not in existing_names:
        existing_names.add(base.lower())
        return base
        
    # Resolve collision
    counter = 1
    while True:
        suffix = f"_{counter}"
        max_base_len = 31 - len(suffix)
        candidate = f"{cleaned[:max_base_len]}{suffix}"
        if candidate.lower() not in existing_names:
            existing_names.add(candidate.lower())
            return candidate
        counter += 1


def export_test_cases_xlsx(db: Session) -> bytes:
    cases = db.query(TestCase).filter(TestCase.deleted_at.is_(None)).order_by(TestCase.updated_at.desc()).all()

    wb = Workbook()
    
    # Group test cases by module name
    from collections import defaultdict
    module_groups = defaultdict(list)
    for tc in cases:
        mod_name = tc.module.strip() if tc.module else "General"
        module_groups[mod_name].append(tc)

    sorted_modules = sorted(module_groups.keys())
    if not sorted_modules:
        sorted_modules = ["General"]
        module_groups["General"] = []

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="595959")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    existing_sheets = set()
    first_sheet = True

    for mod in sorted_modules:
        sheet_title = _sanitize_sheet_name(mod, existing_sheets)
        if first_sheet:
            ws = wb.active
            ws.title = sheet_title
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_title)

        for col_idx, header in enumerate(EXPORT_HEADER, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        for row_idx, tc in enumerate(module_groups[mod], start=2):
            sorted_steps = sorted(tc.steps, key=lambda s: s.step_number)

            step_lines = []
            for s in sorted_steps:
                line = f"{s.step_number}. {s.action or ''}"
                if s.test_data:
                    line += f"\n   Test Data: {s.test_data}"
                if s.expected_result:
                    line += f"\n   Expected: {s.expected_result}"
                step_lines.append(line)
            step_actions = "\n".join(step_lines) if step_lines else None

            row_data = [
                tc.display_id,          # A  TC ID
                tc.title,               # B  Title
                tc.type,                # C  Type
                tc.severity,            # D  Severity
                tc.status,              # E  Status
                tc.description,         # F  Description
                tc.preconditions,       # G  Preconditions
                step_actions,           # H  Step Actions
                tc.expected_result,     # I  Expected Result
                tc.notes,               # J  Notes
                tc.automation_status,   # K  Automation Status
                tc.environment,         # L  Environment
                tc.estimated_time,      # M  Estimated Time
                ";".join(tc.tags) if tc.tags else None,  # N  Tags
                tc.requirement_id,      # O  Requirement ID
                tc.assigned_to_name,    # P  Assigned To
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col_idx, header in enumerate(EXPORT_HEADER, start=1):
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(len(header) + 4, 10)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Helpers shared by export & template
# ---------------------------------------------------------------------------

_REQUIRED_KEYS = {"title", "module", "type", "severity", "expected_result", "step_actions"}

_COL_KEYS = list(COL.keys())

# Enum lists for dropdown data-validation
_DV_MAP: dict[str, list[str]] = {
    COL["type"]:             sorted(VALID_TYPES),
    COL["severity"]:         sorted(VALID_SEVERITIES),
    COL["status"]:           sorted(VALID_STATUSES),
    COL["automation_status"]: sorted(v for v in VALID_AUTOMATION_STATUSES if v),
    COL["environment"]:      sorted(v for v in VALID_ENVIRONMENTS if v),
}


def generate_template_xlsx() -> bytes:
    """Return an XLSX template with coloured headers, dropdowns, and one example row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Authentication"

    # ---- fonts & fills -------------------------------------------------
    hdr_font     = Font(bold=True, color="FFFFFF")
    hdr_fill     = PatternFill(fill_type="solid", fgColor="595959")
    hdr_align    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align   = Alignment(vertical="top", wrap_text=True)

    # ---- write headers -------------------------------------------------
    col_letter_map: dict[str, str] = {}
    for col_idx, (key, header) in enumerate(COL.items(), start=1):
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col_idx)
        col_letter_map[header] = col_letter

        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font    = hdr_font
        cell.fill    = hdr_fill
        cell.alignment = hdr_align

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ---- example row ---------------------------------------------------
    example = [
        "",                                          # A  TC ID  (blank = auto-assign)
        "Verify login with valid credentials",        # B  Title
        "Functional",                                 # C  Type
        "Major",                                      # D  Severity
        "Draft",                                      # E  Status
        "Test that a registered user can log in using correct credentials",  # F  Description
        "User must be registered in the system",      # G  Preconditions
        "1. Open the login page\n2. Enter valid username and password\n3. Click the Login button",  # H  Step Actions
        "User is redirected to dashboard after successful login",  # I  Expected Result
        "",                                           # J  Notes
        "Manual",                                     # K  Automation Status
        "Staging",                                    # L  Environment
        "5 min",                                      # M  Estimated Time
        "auth;login",                                 # N  Tags  (semicolon-separated)
        "REQ-AUTH-001",                               # O  Requirement ID
        "",                                           # P  Assigned To
    ]
    for col_idx, value in enumerate(example, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.alignment = body_align

    ws.row_dimensions[2].height = 60

    # ---- dropdown data validation for rows 2-501 -----------------------
    for header, choices in _DV_MAP.items():
        col_letter = col_letter_map.get(header)
        if not col_letter:
            continue
        formula = '"' + ",".join(choices) + '"'
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid value",
            error=f"Choose one of: {', '.join(choices)}",
        )
        dv.sqref = f"{col_letter}2:{col_letter}501"
        ws.add_data_validation(dv)

    # ---- column widths -------------------------------------------------
    for col_idx, (key, header) in enumerate(COL.items(), start=1):
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(len(header) + 4, 10)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _cell_str(cell) -> str | None:
    if cell is None:
        return None
    value = cell.value
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned if cleaned else None


def parse_xlsx_import(file_bytes: bytes, db: Session) -> ParseResult:
    try:
        wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception:
        return ParseResult(
            total_parsed=0,
            valid_rows=[],
            duplicates=[],
            errors=[ImportError(row=0, field="file", message="Invalid or unreadable XLSX file")],
        )

    valid_rows: list[TestCaseImportRow] = []
    duplicates: list[DuplicateRow] = []
    errors: list[ImportError] = []
    total_parsed_rows = 0

    # Group sheet information for multi-sheet parsing
    sheet_data = []

    # Required columns (excluding Module)
    REQUIRED_COLUMNS = {
        COL["title"],
        COL["type"],
        COL["severity"],
        COL["expected_result"],
        COL["step_actions"],
    }

    for ws in wb.worksheets:
        sheet_name = ws.title
        rows = list(ws.iter_rows())
        if not rows:
            continue
        headers = [_cell_str(cell) for cell in rows[0]]
        headers_set = {h for h in headers if h}

        # Check if this sheet is a test-case sheet by seeing if it has any overlapping columns.
        # Otherwise it could be a guidelines sheet, notes sheet, etc.
        if not headers_set.intersection(REQUIRED_COLUMNS):
            continue

        missing = REQUIRED_COLUMNS - headers_set
        if missing:
            errors.append(ImportError(
                row=0,
                field=f"Sheet '{sheet_name}' Header",
                message=f"Missing required columns: {', '.join(sorted(missing))}"
            ))
            continue

        header_map = {name: idx for idx, name in enumerate(headers) if name}
        data_rows = rows[1:]
        
        sheet_data.append({
            "sheet_name": sheet_name,
            "header_map": header_map,
            "data_rows": data_rows,
        })
        
        total_parsed_rows += len(data_rows)

    if errors:
        wb.close()
        return ParseResult(
            total_parsed=total_parsed_rows,
            valid_rows=[],
            duplicates=[],
            errors=errors,
        )

    if total_parsed_rows > MAX_IMPORT_ROWS:
        wb.close()
        return ParseResult(
            total_parsed=total_parsed_rows,
            valid_rows=[],
            duplicates=[],
            errors=[ImportError(row=0, field="file", message=f"Too many rows across all sheets: {total_parsed_rows}. Maximum allowed is {MAX_IMPORT_ROWS}")],
        )

    # Gather candidate display IDs for duplicate check across all sheets
    candidate_ids = []
    for s_info in sheet_data:
        h_map = s_info["header_map"]
        display_id_col = COL["display_id"]
        if display_id_col in h_map:
            idx = h_map[display_id_col]
            for row in s_info["data_rows"]:
                if idx < len(row):
                    val = _cell_str(row[idx])
                    if val:
                        candidate_ids.append(val)

    existing_ids: set[str] = set()
    existing_map: dict[str, TestCase] = {}
    if candidate_ids:
        db_results = db.query(TestCase).filter(
            TestCase.display_id.in_(candidate_ids),
            TestCase.deleted_at.is_(None),
        ).all()
        existing_map = {tc.display_id: tc for tc in db_results}
        existing_ids = set(existing_map.keys())

    for s_info in sheet_data:
        sheet_name = s_info["sheet_name"]
        header_map = s_info["header_map"]
        data_rows = s_info["data_rows"]

        def get_cell(row, col_name: str) -> str | None:
            idx = header_map.get(col_name)
            if idx is None:
                return None
            return _cell_str(row[idx]) if idx < len(row) else None

        for row_num, row in enumerate(data_rows, start=2):
            # Skip empty rows (where all cells are empty)
            if not any(_cell_str(cell) for cell in row):
                continue

            row_errors: list[ImportError] = []

            title = get_cell(row, COL["title"])
            if not title:
                row_errors.append(ImportError(row=row_num, field="Title", message=f"Required (in sheet '{sheet_name}')"))

            tc_type = get_cell(row, COL["type"])
            if not tc_type:
                row_errors.append(ImportError(row=row_num, field="Type", message=f"Required (in sheet '{sheet_name}')"))
            elif tc_type not in VALID_TYPES:
                row_errors.append(ImportError(row=row_num, field="Type", message=f"Must be one of: {', '.join(sorted(VALID_TYPES))} (in sheet '{sheet_name}')"))

            severity = get_cell(row, COL["severity"])
            if not severity:
                row_errors.append(ImportError(row=row_num, field="Severity", message=f"Required (in sheet '{sheet_name}')"))
            elif severity not in VALID_SEVERITIES:
                row_errors.append(ImportError(row=row_num, field="Severity", message=f"Must be one of: {', '.join(sorted(VALID_SEVERITIES))} (in sheet '{sheet_name}')"))

            expected_result = get_cell(row, COL["expected_result"])
            if not expected_result:
                row_errors.append(ImportError(row=row_num, field="Expected Result", message=f"Required (in sheet '{sheet_name}')"))

            step_actions_raw = get_cell(row, COL["step_actions"])
            if not step_actions_raw:
                row_errors.append(ImportError(row=row_num, field="Step Actions", message=f"At least one step action is required (in sheet '{sheet_name}')"))

            if row_errors:
                errors.extend(row_errors)
                continue

            parsed_steps = _parse_steps_with_details(step_actions_raw)

            steps: list[ImportStepRow] = []
            for ps in parsed_steps:
                if not ps["action"]:
                    continue
                steps.append(ImportStepRow(
                    action=ps["action"],
                    expected_result=ps["expected_result"],
                    test_data=ps["test_data"],
                ))

            status = get_cell(row, COL["status"]) or "Draft"
            if status not in VALID_STATUSES:
                status = "Draft"

            tags_raw = get_cell(row, COL["tags"])
            tags = [t.strip() for t in tags_raw.split(";") if t.strip()] if tags_raw else None

            display_id = get_cell(row, COL["display_id"])

            # Backward compatibility check: if "Module" column exists in this sheet, use it.
            # Otherwise use the sheet name.
            module_col_val = get_cell(row, "Module")
            module = module_col_val if module_col_val else sheet_name

            import_row = TestCaseImportRow(
                row_index=row_num,
                display_id=display_id,
                title=title,
                description=get_cell(row, COL["description"]),
                module=module,
                type=tc_type,
                severity=severity,
                status=status,
                assigned_to_name=get_cell(row, COL["assigned_to_name"]),
                requirement_id=get_cell(row, COL["requirement_id"]),
                estimated_time=get_cell(row, COL["estimated_time"]),
                tags=tags,
                environment=get_cell(row, COL["environment"]),
                automation_status=get_cell(row, COL["automation_status"]),
                preconditions=get_cell(row, COL["preconditions"]),
                expected_result=expected_result,
                notes=get_cell(row, COL["notes"]),
                steps=steps,
            )

            if display_id and display_id in existing_ids:
                existing_tc = existing_map[display_id]
                duplicates.append(DuplicateRow(
                    row_index=row_num,
                    display_id=display_id,
                    import_title=title,
                    existing_title=existing_tc.title,
                    existing_status=existing_tc.status,
                ))
                valid_rows.append(import_row)
            else:
                valid_rows.append(import_row)

    wb.close()

    if not valid_rows and not duplicates and not errors:
        errors.append(ImportError(row=0, field="file", message="No valid sheets containing test cases were found in the uploaded file."))

    return ParseResult(
        total_parsed=total_parsed_rows,
        valid_rows=valid_rows,
        duplicates=duplicates,
        errors=errors,
    )


def _generate_display_id(db: Session) -> str:
    result = db.execute(
        select(func.max(TestCase.display_id)).where(TestCase.display_id.like("CLR-TC-%"))
    ).scalar()
    if result:
        try:
            max_num = int(result.split("-")[-1])
        except (ValueError, IndexError):
            max_num = 0
    else:
        max_num = 0
    return f"CLR-TC-{max_num + 1:03d}"


def _resolve_user_id(db: Session, name: str | None) -> uuid.UUID | None:
    if not name:
        return None
    user = db.query(User).filter(User.name.ilike(name)).first()
    return user.id if user else None


def execute_import(db: Session, request: ImportExecuteRequest) -> ImportResult:
    action_map: dict[str, str] = {da.display_id: da.action for da in request.duplicate_actions}

    created = 0
    skipped = 0
    overwritten = 0
    errors: list[ImportError] = []

    now = datetime.now(timezone.utc)

    for import_row in request.rows:
        display_id = import_row.display_id
        is_duplicate = bool(display_id and display_id in action_map)
        duplicate_action = action_map.get(display_id, "skip") if is_duplicate else None

        if is_duplicate and duplicate_action == "skip":
            skipped += 1
            continue

        try:
            assignee_id = _resolve_user_id(db, import_row.assigned_to_name)

            if is_duplicate and duplicate_action == "overwrite":
                existing = db.query(TestCase).filter(
                    TestCase.display_id == display_id,
                    TestCase.deleted_at.is_(None),
                ).first()

                if not existing:
                    skipped += 1
                    continue

                existing.title = import_row.title
                existing.description = import_row.description
                existing.module = import_row.module
                existing.type = import_row.type
                existing.severity = import_row.severity
                existing.status = import_row.status
                existing.assigned_to = assignee_id
                existing.requirement_id = import_row.requirement_id
                existing.estimated_time = import_row.estimated_time
                existing.tags = import_row.tags
                existing.environment = import_row.environment
                existing.automation_status = import_row.automation_status
                existing.preconditions = import_row.preconditions
                existing.expected_result = import_row.expected_result
                existing.notes = import_row.notes
                existing.updated_at = now

                db.query(TestStep).filter(TestStep.test_case_id == existing.id).delete()

                for step_idx, step in enumerate(import_row.steps):
                    db.add(TestStep(
                        test_case_id=existing.id,
                        step_number=step_idx + 1,
                        action=step.action,
                        expected_result=step.expected_result,
                        test_data=step.test_data,
                        status="Not Run",
                    ))

                db.flush()
                overwritten += 1

            else:
                new_display_id = _generate_display_id(db)
                new_tc = TestCase(
                    display_id=new_display_id,
                    title=import_row.title,
                    description=import_row.description,
                    module=import_row.module,
                    type=import_row.type,
                    severity=import_row.severity,
                    status=import_row.status,
                    assigned_to=assignee_id,
                    requirement_id=import_row.requirement_id,
                    estimated_time=import_row.estimated_time,
                    tags=import_row.tags,
                    environment=import_row.environment,
                    automation_status=import_row.automation_status,
                    preconditions=import_row.preconditions,
                    expected_result=import_row.expected_result,
                    notes=import_row.notes,
                    created_at=now,
                    updated_at=now,
                )
                db.add(new_tc)
                db.flush()

                for step_idx, step in enumerate(import_row.steps):
                    db.add(TestStep(
                        test_case_id=new_tc.id,
                        step_number=step_idx + 1,
                        action=step.action,
                        expected_result=step.expected_result,
                        test_data=step.test_data,
                        status="Not Run",
                    ))

                db.flush()
                created += 1

        except Exception as exc:
            errors.append(ImportError(
                row=import_row.row_index,
                field="__row__",
                message=str(exc),
            ))

    if not errors:
        db.commit()
    else:
        db.rollback()

    return ImportResult(created=created, skipped=skipped, overwritten=overwritten, errors=errors)
